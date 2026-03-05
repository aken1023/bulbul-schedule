import json
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COMPANY_NAME = '愛立康有限公司 統編83460654'
COMPANY_ADDR = '台北市大安區忠孝東路四段 295 號 3 樓'

THIN = Side(style='thin')
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_LEFT = Border(left=Side(style='medium'))
BORDER_RIGHT = Border(right=Side(style='medium'))
BORDER_TOP = Border(top=Side(style='medium'))
BORDER_BOTTOM = Border(bottom=Side(style='medium'))

YELLOW_FILL = PatternFill('solid', fgColor='FFFFEB9C')
GREEN_FILL = PatternFill('solid', fgColor='FFC6EFCE')
HEADER_FILL = PatternFill('solid', fgColor='FFFFF2CC')


def _set_border_range(ws, min_row, max_row, min_col, max_col):
    med = Side(style='medium')
    thin = Side(style='thin')
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            left = med if c == min_col else thin
            right = med if c == max_col else thin
            top = med if r == min_row else thin
            bottom = med if r == max_row else thin
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def _tw_date(d):
    if not d:
        return ''
    return f'{d.year - 1911}.{d.month:02d}.{d.day:02d}'


def _tw_year_month(year, month):
    return f'{year - 1911}.{month:02d}'


def export_salary_slip(employee, salary, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = '薪資明細'

    # Column widths
    col_widths = [2, 6, 22, 4, 12, 4, 18, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row 1-2: Company header
    ws.merge_cells('A1:H1')
    ws['A1'] = COMPANY_NAME
    ws['A1'].font = Font(size=14, bold=True, color='800080')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:H2')
    ws['A2'] = COMPANY_ADDR
    ws['A2'].font = Font(size=11, color='800080')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Row 4: Employee info row 1
    row = 4
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = f'姓名：{employee.name}  {employee.emp_code or ""}'
    ws[f'A{row}'].font = Font(size=11, bold=True)

    ws.merge_cells(f'D{row}:E{row}')
    ws[f'D{row}'] = f'薪資月份:{_tw_year_month(salary["year"], salary["month"])}'
    ws[f'D{row}'].font = Font(size=11)

    ws.merge_cells(f'F{row}:H{row}')
    ws[f'F{row}'] = f'到職日:{_tw_date(employee.start_date)}'
    ws[f'F{row}'].font = Font(size=11)

    for c in range(1, 9):
        ws.cell(row=row, column=c).fill = YELLOW_FILL

    # Row 5: Employee info row 2
    row = 5
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = f'職稱  {employee.job_title or "照服員"}'
    ws[f'A{row}'].font = Font(size=11, bold=True)

    payment_date = salary.get('payment_date')
    pd_str = _tw_date(payment_date) if payment_date else ''
    ws.merge_cells(f'D{row}:E{row}')
    ws[f'D{row}'] = f'入帳日期:{pd_str}'
    ws[f'D{row}'].font = Font(size=11)

    ws.merge_cells(f'F{row}:H{row}')
    bank = employee.bank_account or ''
    ws[f'F{row}'] = f'帳號{bank}' if bank else ''
    ws[f'F{row}'].font = Font(size=11, color='FF0000')

    for c in range(1, 9):
        ws.cell(row=row, column=c).fill = YELLOW_FILL

    _set_border_range(ws, 4, 5, 1, 8)

    # Row 7: Salary detail header
    row = 7
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = '薪資明細表（元）'
    ws[f'A{row}'].font = Font(size=13, bold=True, color='006400')
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    ws[f'A{row}'].fill = HEADER_FILL

    # Build earnings list
    earnings = []
    day_shifts = salary.get('day_shifts', 0)
    night_shifts = salary.get('night_shifts', 0)
    day_rate = salary.get('day_rate', 0)
    night_rate = salary.get('night_rate', 0)

    if day_shifts > 0:
        loc = employee.location or ''
        earnings.append((f'{loc}日班{day_shifts}天({day_rate})', day_shifts * day_rate))
    if night_shifts > 0:
        loc = employee.location or ''
        earnings.append((f'{loc}夜班{night_shifts}天({night_rate})', night_shifts * night_rate))

    extra_earnings = salary.get('extra_earnings', [])
    if isinstance(extra_earnings, str):
        extra_earnings = json.loads(extra_earnings) if extra_earnings else []
    for item in extra_earnings:
        earnings.append((item['name'], item['amount']))

    # Build deductions list
    deductions = [
        ('提早撥款本月薪資', salary.get('advance_pay', 0)),
        ('健保費眷屬', salary.get('health_insurance', 0)),
        ('福利金', salary.get('welfare_fund', 0)),
        ('請假', salary.get('leave_deduct', 0)),
        ('匯費', salary.get('transfer_fee', 0)),
    ]
    extra_deductions = salary.get('extra_deductions', [])
    if isinstance(extra_deductions, str):
        extra_deductions = json.loads(extra_deductions) if extra_deductions else []
    for item in extra_deductions:
        deductions.append((item['name'], item['amount']))

    detail_rows = max(len(earnings), len(deductions), 8)

    # Detail section
    start_row = 8
    for i in range(detail_rows):
        r = start_row + i
        # Left: 發金額
        if i == 0:
            ws.cell(row=r, column=1).value = '發'
        elif i == 1:
            ws.cell(row=r, column=1).value = '金'
        elif i == 2:
            ws.cell(row=r, column=1).value = '額'
        ws.cell(row=r, column=1).fill = GREEN_FILL
        ws.cell(row=r, column=1).font = Font(size=11, bold=True)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')

        if i < len(earnings):
            ws.merge_cells(f'B{r}:D{r}')
            ws.cell(row=r, column=2).value = earnings[i][0]
            ws.cell(row=r, column=2).fill = GREEN_FILL
            ws.cell(row=r, column=5).value = earnings[i][1]
            ws.cell(row=r, column=5).number_format = '#,##0'
        else:
            ws.merge_cells(f'B{r}:D{r}')
            ws.cell(row=r, column=2).fill = GREEN_FILL

        ws.cell(row=r, column=5).fill = GREEN_FILL
        ws.cell(row=r, column=5).alignment = Alignment(horizontal='right')

        # Middle separator
        if i == 0:
            ws.cell(row=r, column=6).value = '扣'
        elif i == 1:
            ws.cell(row=r, column=6).value = '金'
        elif i == 2:
            ws.cell(row=r, column=6).value = '額'
        ws.cell(row=r, column=6).font = Font(size=11, bold=True)
        ws.cell(row=r, column=6).alignment = Alignment(horizontal='center', vertical='center')

        # Right: 扣金額
        if i < len(deductions):
            ws.cell(row=r, column=7).value = deductions[i][0]
            ws.cell(row=r, column=8).value = deductions[i][1]
            ws.cell(row=r, column=8).number_format = '#,##0'
        ws.cell(row=r, column=8).alignment = Alignment(horizontal='right')

    # Subtotal row
    sub_row = start_row + detail_rows
    ws.cell(row=sub_row, column=1).fill = GREEN_FILL
    ws.merge_cells(f'B{sub_row}:D{sub_row}')
    ws.cell(row=sub_row, column=2).value = '小計'
    ws.cell(row=sub_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=sub_row, column=2).fill = GREEN_FILL

    earn_total = sum(e[1] for e in earnings)
    ws.cell(row=sub_row, column=5).value = earn_total
    ws.cell(row=sub_row, column=5).number_format = '#,##0'
    ws.cell(row=sub_row, column=5).font = Font(size=11, bold=True)
    ws.cell(row=sub_row, column=5).fill = GREEN_FILL
    ws.cell(row=sub_row, column=5).alignment = Alignment(horizontal='right')

    ws.cell(row=sub_row, column=7).value = '小計'
    ws.cell(row=sub_row, column=7).font = Font(size=11, bold=True)

    deduct_total = sum(d[1] for d in deductions)
    ws.cell(row=sub_row, column=8).value = deduct_total
    ws.cell(row=sub_row, column=8).number_format = '#,##0'
    ws.cell(row=sub_row, column=8).font = Font(size=11, bold=True)
    ws.cell(row=sub_row, column=8).alignment = Alignment(horizontal='right')

    _set_border_range(ws, 7, sub_row, 1, 8)

    # Note row
    note_row = sub_row + 1
    ws.merge_cells(f'A{note_row}:C{note_row}')
    ws[f'A{note_row}'] = '備註'
    ws[f'A{note_row}'].font = Font(size=11, bold=True)
    ws.merge_cells(f'D{note_row}:H{note_row}')
    ws[f'D{note_row}'] = salary.get('note', '')
    _set_border_range(ws, note_row, note_row, 1, 8)

    # Actual pay row
    pay_row = note_row + 1
    ws.merge_cells(f'A{pay_row}:C{pay_row}')
    ws[f'A{pay_row}'] = '實發金額（元）'
    ws[f'A{pay_row}'].font = Font(size=12, bold=True)
    ws[f'A{pay_row}'].fill = YELLOW_FILL

    actual_pay = earn_total - deduct_total
    ws.merge_cells(f'D{pay_row}:G{pay_row}')
    ws[f'D{pay_row}'].fill = PatternFill('solid', fgColor='FFFF0000')

    ws[f'H{pay_row}'] = actual_pay
    ws[f'H{pay_row}'].number_format = '#,##0'
    ws[f'H{pay_row}'].font = Font(size=14, bold=True)
    ws[f'H{pay_row}'].alignment = Alignment(horizontal='right')
    _set_border_range(ws, pay_row, pay_row, 1, 8)

    # Footer notes
    fn_row = pay_row + 2
    ws[f'B{fn_row}'] = '備註：'
    ws[f'B{fn_row}'].font = Font(size=10, color='0000FF')
    fn_row += 1
    ws[f'B{fn_row}'] = '1乙若對本款項有問題請於24小時內提出，逾期恕不受理 謝謝!'
    ws[f'B{fn_row}'].font = Font(size=10, color='0000FF')
    fn_row += 1
    ws[f'B{fn_row}'] = '2.工作滿3個月者如非台灣銀行帳戶須自付30元轉帳費'
    ws[f'B{fn_row}'].font = Font(size=10, color='0000FF')
    fn_row += 1
    ws[f'B{fn_row}'] = '3.富邦照服員責任險生效日:115.01.23-116.01.22'
    ws[f'B{fn_row}'].font = Font(size=10, color='0000FF')

    # Date footer
    fn_row += 2
    tw_year = salary['year'] - 1911
    ws.merge_cells(f'C{fn_row}:H{fn_row}')
    ws[f'C{fn_row}'] = f'{tw_year}     年     {salary["month"]}月     {date.today().day} 日'
    ws[f'C{fn_row}'].font = Font(size=16, bold=True)
    ws[f'C{fn_row}'].alignment = Alignment(horizontal='center')

    # Print settings
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    wb.save(output_path)
    return output_path
