import os
import json
import calendar
from datetime import date, datetime
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for
from models import db, Employee, Schedule, SalaryRecord, Location, EmployeePreference, EmployeeTimeOff, StaffingRequirement, EmployeeRateHistory
from export import export_salary_slip

app = Flask(__name__)
DATABASE_URI = os.environ.get(
    'DATABASE_URL',
    'mysql+pymysql://A999:1023@122.100.99.161:43306/bulbul_schedule?charset=utf8mb4'
)
app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URI
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'dev-secret-key'

db.init_app(app)

with app.app_context():
    db.create_all()
    # Ensure default locations exist
    for loc_name in ['雙和', '北醫']:
        if not Location.query.filter_by(name=loc_name).first():
            db.session.add(Location(name=loc_name))
    db.session.commit()

import tempfile
EXPORT_DIR = os.path.join(tempfile.gettempdir(), 'exports')
os.makedirs(EXPORT_DIR, exist_ok=True)


def _parse_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except ValueError:
        return None


def _get_effective_rate(emp, year, month):
    """取得員工在指定年月適用的日班/夜班費率（依生效日期查找最近一筆）"""
    target = date(year, month, 1)
    rate = EmployeeRateHistory.query.filter(
        EmployeeRateHistory.employee_id == emp.id,
        EmployeeRateHistory.effective_date <= target
    ).order_by(EmployeeRateHistory.effective_date.desc()).first()
    if rate:
        return rate.day_rate, rate.night_rate
    return emp.day_rate, emp.night_rate


# --- Pages ---

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/employees')
def employees_page():
    emps = Employee.query.filter_by(is_active=True).order_by(Employee.id).all()
    return render_template('employees.html', employees=emps)


@app.route('/schedule')
def schedule_page():
    today = date.today()
    prev_month = today.month - 1 if today.month > 1 else 12
    prev_year = today.year if today.month > 1 else today.year - 1
    year = request.args.get('year', prev_year, type=int)
    month = request.args.get('month', prev_month, type=int)
    location = request.args.get('location', '')
    return render_template('schedule.html', year=year, month=month, location=location)


@app.route('/salary')
def salary_page():
    today = date.today()
    prev_month = today.month - 1 if today.month > 1 else 12
    prev_year = today.year if today.month > 1 else today.year - 1
    year = request.args.get('year', prev_year, type=int)
    month = request.args.get('month', prev_month, type=int)
    return render_template('salary.html', year=year, month=month)


@app.route('/stats')
def stats_page():
    today = date.today()
    prev_month = today.month - 1 if today.month > 1 else 12
    prev_year = today.year if today.month > 1 else today.year - 1
    year = request.args.get('year', prev_year, type=int)
    month = request.args.get('month', prev_month, type=int)
    return render_template('stats.html', year=year, month=month)


# --- API: Locations ---

@app.route('/api/locations', methods=['GET'])
def api_locations():
    locs = Location.query.filter_by(is_active=True).order_by(Location.id).all()
    return jsonify([l.to_dict() for l in locs])


@app.route('/api/locations', methods=['POST'])
def api_create_location():
    data = request.json
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': '地點名稱不可為空'}), 400
    existing = Location.query.filter_by(name=name).first()
    if existing:
        if not existing.is_active:
            existing.is_active = True
            db.session.commit()
            return jsonify(existing.to_dict()), 200
        return jsonify({'error': '地點已存在'}), 409
    loc = Location(name=name)
    db.session.add(loc)
    db.session.commit()
    return jsonify(loc.to_dict()), 201


@app.route('/api/locations/<int:lid>', methods=['PUT'])
def api_update_location(lid):
    loc = Location.query.get_or_404(lid)
    data = request.json
    new_name = data.get('name', '').strip()
    if not new_name:
        return jsonify({'error': '地點名稱不可為空'}), 400
    old_name = loc.name
    loc.name = new_name
    if 'max_consecutive_days' in data:
        loc.max_consecutive_days = int(data['max_consecutive_days'])
    # Update all employees with the old location name
    Employee.query.filter_by(location=old_name).update({'location': new_name})
    db.session.commit()
    return jsonify(loc.to_dict())


@app.route('/api/locations/<int:lid>', methods=['DELETE'])
def api_delete_location(lid):
    loc = Location.query.get_or_404(lid)
    loc.is_active = False
    db.session.commit()
    return jsonify({'ok': True})


# --- API: Employees ---

@app.route('/api/employees', methods=['GET'])
def api_employees():
    q = Employee.query.filter_by(is_active=True)
    loc = request.args.get('location')
    if loc:
        q = q.filter_by(location=loc)
    return jsonify([e.to_dict() for e in q.order_by(Employee.id).all()])


@app.route('/api/employees', methods=['POST'])
def api_create_employee():
    data = request.json
    emp = Employee(
        name=data['name'],
        emp_code=data.get('emp_code', ''),
        start_date=_parse_date(data.get('start_date')),
        location=data.get('location', ''),
        job_title=data.get('job_title', '照服員'),
        bank_account=data.get('bank_account', ''),
        day_rate=int(data.get('day_rate', 3068)),
        night_rate=int(data.get('night_rate', 3168)),
    )
    db.session.add(emp)
    db.session.commit()
    return jsonify(emp.to_dict()), 201


@app.route('/api/employees/<int:eid>', methods=['PUT'])
def api_update_employee(eid):
    emp = Employee.query.get_or_404(eid)
    data = request.json
    for key in ['name', 'emp_code', 'location', 'job_title', 'bank_account']:
        if key in data:
            setattr(emp, key, data[key])
    if 'start_date' in data:
        emp.start_date = _parse_date(data['start_date'])
    if 'day_rate' in data:
        emp.day_rate = int(data['day_rate'])
    if 'night_rate' in data:
        emp.night_rate = int(data['night_rate'])
    db.session.commit()
    return jsonify(emp.to_dict())


@app.route('/api/employees/<int:eid>', methods=['DELETE'])
def api_delete_employee(eid):
    emp = Employee.query.get_or_404(eid)
    emp.is_active = False
    db.session.commit()
    return jsonify({'ok': True})


# --- API: Employee Rate History (薪資費率歷史) ---

@app.route('/api/rate-history/<int:eid>', methods=['GET'])
def api_get_rate_history(eid):
    Employee.query.get_or_404(eid)
    records = EmployeeRateHistory.query.filter_by(employee_id=eid)\
        .order_by(EmployeeRateHistory.effective_date).all()
    return jsonify([r.to_dict() for r in records])


@app.route('/api/rate-history/<int:eid>', methods=['POST'])
def api_add_rate_history(eid):
    Employee.query.get_or_404(eid)
    data = request.json
    d = _parse_date(data.get('effective_date'))
    if not d:
        return jsonify({'error': '日期格式錯誤'}), 400
    existing = EmployeeRateHistory.query.filter_by(employee_id=eid, effective_date=d).first()
    if existing:
        existing.day_rate = int(data['day_rate'])
        existing.night_rate = int(data['night_rate'])
        db.session.commit()
        return jsonify(existing.to_dict())
    r = EmployeeRateHistory(
        employee_id=eid, effective_date=d,
        day_rate=int(data['day_rate']), night_rate=int(data['night_rate'])
    )
    db.session.add(r)
    db.session.commit()
    return jsonify(r.to_dict()), 201


@app.route('/api/rate-history/<int:eid>/<int:rid>', methods=['DELETE'])
def api_delete_rate_history(eid, rid):
    r = EmployeeRateHistory.query.get_or_404(rid)
    db.session.delete(r)
    db.session.commit()
    return jsonify({'ok': True})


# --- API: Employee Time Off (預定排休) ---

@app.route('/api/time-off/<int:eid>', methods=['GET'])
def api_get_time_off(eid):
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    q = EmployeeTimeOff.query.filter_by(employee_id=eid)
    if year and month:
        _, dim = calendar.monthrange(year, month)
        q = q.filter(EmployeeTimeOff.date >= date(year, month, 1),
                      EmployeeTimeOff.date <= date(year, month, dim))
    return jsonify([t.to_dict() for t in q.order_by(EmployeeTimeOff.date).all()])


@app.route('/api/time-off/<int:eid>', methods=['POST'])
def api_add_time_off(eid):
    Employee.query.get_or_404(eid)
    data = request.json
    d = _parse_date(data['date'])
    if not d:
        return jsonify({'error': '日期格式錯誤'}), 400
    existing = EmployeeTimeOff.query.filter_by(employee_id=eid, date=d).first()
    if existing:
        existing.reason = data.get('reason', '')
        db.session.commit()
        return jsonify(existing.to_dict())
    t = EmployeeTimeOff(employee_id=eid, date=d, reason=data.get('reason', ''))
    db.session.add(t)
    db.session.commit()
    return jsonify(t.to_dict()), 201


@app.route('/api/time-off/<int:eid>/<int:tid>', methods=['DELETE'])
def api_delete_time_off(eid, tid):
    t = EmployeeTimeOff.query.get_or_404(tid)
    db.session.delete(t)
    db.session.commit()
    return jsonify({'ok': True})


# --- API: Employee Preferences ---

@app.route('/api/employee-preferences/<int:eid>', methods=['GET'])
def api_get_preference(eid):
    Employee.query.get_or_404(eid)
    pref = EmployeePreference.query.filter_by(employee_id=eid).first()
    if not pref:
        return jsonify({'employee_id': eid, 'allowed_locations': [], 'allowed_shifts': 'both', 'min_days': 0, 'max_days': 31})
    return jsonify(pref.to_dict())


@app.route('/api/employee-preferences/<int:eid>', methods=['PUT'])
def api_set_preference(eid):
    Employee.query.get_or_404(eid)
    data = request.json
    pref = EmployeePreference.query.filter_by(employee_id=eid).first()
    if not pref:
        pref = EmployeePreference(employee_id=eid)
        db.session.add(pref)
    pref.allowed_locations = json.dumps(data.get('allowed_locations', []), ensure_ascii=False)
    pref.allowed_shifts = data.get('allowed_shifts', 'both')
    pref.min_days = data.get('min_days', 0)
    pref.max_days = data.get('max_days', 31)
    pref.schedule_mode = data.get('schedule_mode', 'none')
    pref.work_weekdays = json.dumps(data.get('work_weekdays', []))
    pref.pattern_work = data.get('pattern_work', 5)
    pref.pattern_off = data.get('pattern_off', 2)
    db.session.commit()
    return jsonify(pref.to_dict())


@app.route('/api/employee-preferences', methods=['GET'])
def api_all_preferences():
    prefs = EmployeePreference.query.all()
    pref_map = {p.employee_id: p.to_dict() for p in prefs}
    employees = Employee.query.filter_by(is_active=True).order_by(Employee.id).all()
    result = []
    for emp in employees:
        p = pref_map.get(emp.id, {
            'employee_id': emp.id, 'allowed_locations': [],
            'allowed_shifts': 'both', 'min_days': 0, 'max_days': 31
        })
        p['employee_name'] = emp.name
        p['employee_location'] = emp.location
        result.append(p)
    return jsonify(result)


# --- API: Schedule ---

@app.route('/api/schedule', methods=['GET'])
def api_get_schedule():
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    location = request.args.get('location', '')

    _, days_in_month = calendar.monthrange(year, month)
    start = date(year, month, 1)
    end = date(year, month, days_in_month)

    q = Employee.query.filter_by(is_active=True)
    if location:
        q = q.filter_by(location=location)
    employees = q.order_by(Employee.id).all()

    emp_ids = [e.id for e in employees]
    schedules = Schedule.query.filter(
        Schedule.employee_id.in_(emp_ids),
        Schedule.date >= start,
        Schedule.date <= end
    ).all()

    sched_map = {}
    for s in schedules:
        sched_map[(s.employee_id, s.date.day)] = s.shift

    result = []
    for emp in employees:
        shifts = {}
        for d in range(1, days_in_month + 1):
            shifts[str(d)] = sched_map.get((emp.id, d), 'OFF')
        result.append({
            'employee': emp.to_dict(),
            'shifts': shifts,
        })

    return jsonify({
        'year': year,
        'month': month,
        'days_in_month': days_in_month,
        'data': result,
    })


@app.route('/api/schedule', methods=['POST'])
def api_set_schedule():
    data = request.json
    emp_id = data['employee_id']
    year = data['year']
    month = data['month']
    day = data['day']
    shift = data['shift']

    d = date(year, month, day)
    existing = Schedule.query.filter_by(employee_id=emp_id, date=d).first()
    if existing:
        existing.shift = shift
    else:
        db.session.add(Schedule(employee_id=emp_id, date=d, shift=shift))
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/schedule/batch', methods=['POST'])
def api_batch_schedule():
    data = request.json
    emp_id = data['employee_id']
    year = data['year']
    month = data['month']
    shifts = data['shifts']  # {"1":"D","2":"N",...}

    for day_str, shift in shifts.items():
        d = date(year, month, int(day_str))
        existing = Schedule.query.filter_by(employee_id=emp_id, date=d).first()
        if existing:
            existing.shift = shift
        else:
            db.session.add(Schedule(employee_id=emp_id, date=d, shift=shift))
    db.session.commit()
    return jsonify({'ok': True})


# --- API: Salary ---

@app.route('/api/salary/calculate', methods=['GET'])
def api_calculate_salary():
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    _, days_in_month = calendar.monthrange(year, month)
    start = date(year, month, 1)
    end = date(year, month, days_in_month)

    employees = Employee.query.filter_by(is_active=True).order_by(Employee.id).all()
    results = []

    for emp in employees:
        schedules = Schedule.query.filter(
            Schedule.employee_id == emp.id,
            Schedule.date >= start,
            Schedule.date <= end
        ).all()

        sched_day_count = sum(1 for s in schedules if s.shift == 'D')
        sched_night_count = sum(1 for s in schedules if s.shift == 'N')

        # Check for existing salary record (overrides auto-calc)
        rec = SalaryRecord.query.filter_by(
            employee_id=emp.id, year=year, month=month
        ).first()

        extra_earnings = []
        deductions = {
            'advance_pay': 0, 'health_insurance': 0,
            'welfare_fund': 0, 'leave_deduct': 0, 'transfer_fee': 0,
        }
        extra_deductions = []
        note = ''
        payment_date = None

        if rec:
            # Use saved record's rates and counts (may differ from schedule)
            day_count = rec.day_shifts
            night_count = rec.night_shifts
            day_rate = rec.day_rate
            night_rate = rec.night_rate
            extra_earnings = json.loads(rec.extra_earnings) if rec.extra_earnings else []
            deductions['advance_pay'] = rec.advance_pay
            deductions['health_insurance'] = rec.health_insurance
            deductions['welfare_fund'] = rec.welfare_fund
            deductions['leave_deduct'] = rec.leave_deduct
            deductions['transfer_fee'] = rec.transfer_fee
            extra_deductions = json.loads(rec.extra_deductions) if rec.extra_deductions else []
            note = rec.note or ''
            payment_date = rec.payment_date.isoformat() if rec.payment_date else None
        else:
            # No saved record, auto-calc from schedule + effective rate
            day_count = sched_day_count
            night_count = sched_night_count
            day_rate, night_rate = _get_effective_rate(emp, year, month)

        off_count = days_in_month - sched_day_count - sched_night_count
        day_total = day_count * day_rate
        night_total = night_count * night_rate
        gross = day_total + night_total

        extra_earn_total = sum(e['amount'] for e in extra_earnings)
        extra_deduct_total = sum(d['amount'] for d in extra_deductions)
        deduct_total = sum(deductions.values()) + extra_deduct_total
        actual_pay = gross + extra_earn_total - deduct_total

        results.append({
            'employee': emp.to_dict(),
            'day_shifts': day_count,
            'night_shifts': night_count,
            'off_days': off_count,
            'day_rate': day_rate,
            'night_rate': night_rate,
            'day_total': day_total,
            'night_total': night_total,
            'gross': gross,
            'extra_earnings': extra_earnings,
            'extra_earn_total': extra_earn_total,
            'deductions': deductions,
            'extra_deductions': extra_deductions,
            'deduct_total': deduct_total,
            'actual_pay': actual_pay,
            'note': note,
            'payment_date': payment_date,
        })

    return jsonify({'year': year, 'month': month, 'data': results})


@app.route('/api/salary/save', methods=['POST'])
def api_save_salary():
    data = request.json
    emp_id = data['employee_id']
    year = data['year']
    month = data['month']

    rec = SalaryRecord.query.filter_by(
        employee_id=emp_id, year=year, month=month
    ).first()
    if not rec:
        rec = SalaryRecord(employee_id=emp_id, year=year, month=month)
        db.session.add(rec)

    rec.day_shifts = data.get('day_shifts', 0)
    rec.night_shifts = data.get('night_shifts', 0)
    rec.day_rate = data.get('day_rate', 0)
    rec.night_rate = data.get('night_rate', 0)
    rec.extra_earnings = json.dumps(data.get('extra_earnings', []), ensure_ascii=False)
    rec.advance_pay = data.get('advance_pay', 0)
    rec.health_insurance = data.get('health_insurance', 0)
    rec.welfare_fund = data.get('welfare_fund', 0)
    rec.leave_deduct = data.get('leave_deduct', 0)
    rec.transfer_fee = data.get('transfer_fee', 0)
    rec.extra_deductions = json.dumps(data.get('extra_deductions', []), ensure_ascii=False)
    rec.note = data.get('note', '')
    rec.payment_date = _parse_date(data.get('payment_date'))

    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/salary/export/<int:emp_id>', methods=['GET'])
def api_export_salary(emp_id):
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    emp = Employee.query.get_or_404(emp_id)

    _, days_in_month = calendar.monthrange(year, month)
    start = date(year, month, 1)
    end = date(year, month, days_in_month)

    schedules = Schedule.query.filter(
        Schedule.employee_id == emp.id,
        Schedule.date >= start,
        Schedule.date <= end
    ).all()

    day_count = sum(1 for s in schedules if s.shift == 'D')
    night_count = sum(1 for s in schedules if s.shift == 'N')

    rec = SalaryRecord.query.filter_by(
        employee_id=emp.id, year=year, month=month
    ).first()

    eff_day, eff_night = _get_effective_rate(emp, year, month)
    salary_data = {
        'year': year,
        'month': month,
        'day_shifts': day_count,
        'night_shifts': night_count,
        'day_rate': rec.day_rate if rec else eff_day,
        'night_rate': rec.night_rate if rec else eff_night,
        'extra_earnings': json.loads(rec.extra_earnings) if rec and rec.extra_earnings else [],
        'advance_pay': rec.advance_pay if rec else 0,
        'health_insurance': rec.health_insurance if rec else 0,
        'welfare_fund': rec.welfare_fund if rec else 0,
        'leave_deduct': rec.leave_deduct if rec else 0,
        'transfer_fee': rec.transfer_fee if rec else 0,
        'extra_deductions': json.loads(rec.extra_deductions) if rec and rec.extra_deductions else [],
        'note': rec.note if rec else '',
        'payment_date': rec.payment_date if rec else None,
    }

    tw_ym = f'{year - 1911}{month:02d}'
    filename = f'薪資明細_{emp.name}_{tw_ym}.xlsx'
    filepath = os.path.join(EXPORT_DIR, filename)
    export_salary_slip(emp, salary_data, filepath)

    return send_file(filepath, as_attachment=True, download_name=filename)


@app.route('/api/salary/export-all', methods=['GET'])
def api_export_all():
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    employees = Employee.query.filter_by(is_active=True).order_by(Employee.id).all()
    _, days_in_month = calendar.monthrange(year, month)
    start = date(year, month, 1)
    end = date(year, month, days_in_month)

    for emp in employees:
        schedules = Schedule.query.filter(
            Schedule.employee_id == emp.id,
            Schedule.date >= start,
            Schedule.date <= end
        ).all()

        day_count = sum(1 for s in schedules if s.shift == 'D')
        night_count = sum(1 for s in schedules if s.shift == 'N')

        rec = SalaryRecord.query.filter_by(
            employee_id=emp.id, year=year, month=month
        ).first()

        eff_day, eff_night = _get_effective_rate(emp, year, month)
        salary_data = {
            'year': year, 'month': month,
            'day_shifts': day_count, 'night_shifts': night_count,
            'day_rate': rec.day_rate if rec else eff_day,
            'night_rate': rec.night_rate if rec else eff_night,
            'extra_earnings': json.loads(rec.extra_earnings) if rec and rec.extra_earnings else [],
            'advance_pay': rec.advance_pay if rec else 0,
            'health_insurance': rec.health_insurance if rec else 0,
            'welfare_fund': rec.welfare_fund if rec else 0,
            'leave_deduct': rec.leave_deduct if rec else 0,
            'transfer_fee': rec.transfer_fee if rec else 0,
            'extra_deductions': json.loads(rec.extra_deductions) if rec and rec.extra_deductions else [],
            'note': rec.note if rec else '',
            'payment_date': rec.payment_date if rec else None,
        }

        # Export individual slip then copy sheet
        tmp_path = os.path.join(EXPORT_DIR, f'_tmp_{emp.id}.xlsx')
        export_salary_slip(emp, salary_data, tmp_path)

        from openpyxl import load_workbook
        tmp_wb = load_workbook(tmp_path)
        tmp_ws = tmp_wb.active

        new_ws = wb.create_sheet(title=emp.name[:31])
        for row in tmp_ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.alignment = cell.alignment.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.number_format = cell.number_format

        for col_letter, dim in tmp_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = dim.width

        for key in tmp_ws.merged_cells.ranges:
            new_ws.merge_cells(str(key))

        tmp_wb.close()
        try:
            os.remove(tmp_path)
        except PermissionError:
            pass

    tw_ym = f'{year - 1911}{month:02d}'
    filename = f'薪資明細_全體_{tw_ym}.xlsx'
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)

    return send_file(filepath, as_attachment=True, download_name=filename)


# --- API: Staffing Requirements ---

@app.route('/api/staffing-requirements', methods=['GET'])
def api_get_staffing():
    reqs = StaffingRequirement.query.all()
    return jsonify([r.to_dict() for r in reqs])


@app.route('/api/staffing-requirements', methods=['PUT'])
def api_set_staffing():
    data = request.json  # list of {location, shift, count}
    for item in data:
        req = StaffingRequirement.query.filter_by(
            location=item['location'], shift=item['shift']
        ).first()
        if req:
            req.count = item['count']
        else:
            db.session.add(StaffingRequirement(
                location=item['location'], shift=item['shift'], count=item['count']
            ))
    db.session.commit()
    return jsonify({'ok': True})


# --- API: Auto Schedule ---

@app.route('/api/schedule/auto-generate', methods=['POST'])
def api_auto_generate():
    from auto_schedule import auto_generate
    data = request.json
    year = data['year']
    month = data['month']

    # Block past months
    today = date.today()
    if year < today.year or (year == today.year and month < today.month):
        return jsonify({'error': '無法對已過去的月份使用自動排班'}), 400

    employees = Employee.query.filter_by(is_active=True).order_by(Employee.id).all()
    locs = Location.query.filter_by(is_active=True).all()
    loc_names = [l.name for l in locs]

    # Build preferences dict
    prefs_raw = EmployeePreference.query.all()
    preferences = {}
    for p in prefs_raw:
        preferences[p.employee_id] = p.to_dict()

    # Build staffing requirements
    staffing_reqs = {}
    reqs = StaffingRequirement.query.all()
    for r in reqs:
        staffing_reqs[(r.location, r.shift)] = r.count

    loc_max_consecutive = {l.name: (l.max_consecutive_days or 6) for l in locs}

    # Load time-off dates
    _, dim = calendar.monthrange(year, month)
    time_offs = EmployeeTimeOff.query.filter(
        EmployeeTimeOff.date >= date(year, month, 1),
        EmployeeTimeOff.date <= date(year, month, dim)
    ).all()
    time_off_dates = {}
    for t in time_offs:
        time_off_dates.setdefault(t.employee_id, set()).add(t.date.day)

    result, shortages = auto_generate(year, month, employees, preferences, staffing_reqs, loc_names, loc_max_consecutive, time_off_dates)

    # Convert to serializable format
    preview = []
    for emp in employees:
        shifts = {str(d): result[emp.id][d] for d in result[emp.id]}
        preview.append({'employee': emp.to_dict(), 'shifts': shifts})

    return jsonify({'year': year, 'month': month, 'data': preview, 'shortages': shortages})


@app.route('/api/schedule/auto-confirm', methods=['POST'])
def api_auto_confirm():
    data = request.json
    year = data['year']
    month = data['month']
    schedule_data = data['data']  # list of {employee_id, shifts: {day: shift}}

    for item in schedule_data:
        emp_id = item['employee_id']
        for day_str, shift in item['shifts'].items():
            d = date(year, month, int(day_str))
            existing = Schedule.query.filter_by(employee_id=emp_id, date=d).first()
            if existing:
                existing.shift = shift
            else:
                db.session.add(Schedule(employee_id=emp_id, date=d, shift=shift))
    db.session.commit()
    return jsonify({'ok': True})


# --- API: Stats ---

@app.route('/api/stats', methods=['GET'])
def api_stats():
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    _, days_in_month = calendar.monthrange(year, month)
    start = date(year, month, 1)
    end = date(year, month, days_in_month)

    employees = Employee.query.filter_by(is_active=True).order_by(Employee.id).all()
    locs = Location.query.filter_by(is_active=True).order_by(Location.id).all()
    loc_names = [l.name for l in locs]

    # Per-employee stats
    emp_stats = []
    # Per-location per-day stats
    loc_day_d = {l: [0] * days_in_month for l in loc_names}
    loc_day_n = {l: [0] * days_in_month for l in loc_names}
    # Location totals
    loc_totals = {l: {'D': 0, 'N': 0, 'OFF': 0} for l in loc_names}

    for emp in employees:
        schedules = Schedule.query.filter(
            Schedule.employee_id == emp.id,
            Schedule.date >= start, Schedule.date <= end
        ).all()
        sched_map = {s.date.day: s.shift for s in schedules}

        d_count = sum(1 for s in schedules if s.shift == 'D')
        n_count = sum(1 for s in schedules if s.shift == 'N')
        off_count = days_in_month - d_count - n_count

        # Salary
        rec = SalaryRecord.query.filter_by(employee_id=emp.id, year=year, month=month).first()
        if rec:
            day_rate = rec.day_rate
            night_rate = rec.night_rate
            d_pay_count = rec.day_shifts
            n_pay_count = rec.night_shifts
            extra = sum(e['amount'] for e in (json.loads(rec.extra_earnings) if rec.extra_earnings else []))
            deduct = rec.advance_pay + rec.health_insurance + rec.welfare_fund + rec.leave_deduct + rec.transfer_fee
            deduct += sum(d['amount'] for d in (json.loads(rec.extra_deductions) if rec.extra_deductions else []))
        else:
            day_rate, night_rate = _get_effective_rate(emp, year, month)
            d_pay_count = d_count
            n_pay_count = n_count
            extra = 0
            deduct = 0

        actual_pay = d_pay_count * day_rate + n_pay_count * night_rate + extra - deduct

        emp_stats.append({
            'name': emp.name, 'location': emp.location,
            'D': d_count, 'N': n_count, 'OFF': off_count,
            'actual_pay': actual_pay,
        })

        loc = emp.location
        if loc in loc_totals:
            loc_totals[loc]['D'] += d_count
            loc_totals[loc]['N'] += n_count
            loc_totals[loc]['OFF'] += off_count
            for day in range(1, days_in_month + 1):
                shift = sched_map.get(day, 'OFF')
                if shift == 'D':
                    loc_day_d[loc][day - 1] += 1
                elif shift == 'N':
                    loc_day_n[loc][day - 1] += 1

    return jsonify({
        'year': year, 'month': month, 'days_in_month': days_in_month,
        'locations': loc_names,
        'location_totals': loc_totals,
        'location_daily_d': loc_day_d,
        'location_daily_n': loc_day_n,
        'employees': emp_stats,
    })


if __name__ == '__main__':
    app.run(debug=True, port=5000)
