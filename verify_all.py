import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from app import app

errors = 0

def check(label, ok):
    global errors
    if ok:
        print(f'  PASS {label}')
    else:
        print(f'  FAIL {label}')
        errors += 1

print('=' * 60)
print('STEP 1: 頁面載入')
print('=' * 60)
with app.test_client() as c:
    for path in ['/', '/employees', '/schedule?year=2026&month=2', '/salary?year=2026&month=2']:
        r = c.get(path)
        check(f'{path} -> {r.status_code}', r.status_code == 200)

print()
print('=' * 60)
print('STEP 2: 員工資料 (19人)')
print('=' * 60)
with app.test_client() as c:
    emps = c.get('/api/employees').get_json()
    check(f'員工總數={len(emps)}', len(emps) == 19)
    shuanghe = [e for e in emps if e['location'] == '雙和']
    beiyi = [e for e in emps if e['location'] == '北醫']
    check(f'雙和={len(shuanghe)}人', len(shuanghe) == 10)
    check(f'北醫={len(beiyi)}人', len(beiyi) == 9)

print()
print('=' * 60)
print('STEP 3: 排班表 逐日比對')
print('=' * 60)

def normalize(v):
    if not v: return 'OFF'
    v = str(v).strip().upper()
    if v in ('OFF', ''): return 'OFF'
    if v.startswith('D') or v in ('6HR',): return 'D'
    if v.startswith('N'): return 'N'
    return 'OFF'

wb = openpyxl.load_workbook('115.02班aken.xlsx')
ws1 = wb[wb.sheetnames[0]]
ws2 = wb[wb.sheetnames[1]]

excel_shifts = {}
for name, row in {'施秉綬':7,'陳惠美':10,'王明珠':13,'何金艷':16,'全蕙萍':21,'衡怡紅':24,'藍茂禎':27,'林玉燕':30,'邱素蘭':33}.items():
    for col in range(7, 35):
        excel_shifts[(name, col-6)] = normalize(ws1.cell(row=row, column=col).value)

for name, row in {'林淑真':6,'歐南蘭':9,'沈佳榕':12,'鍾宏枝':15,'張瀞文':18,'王柏欣':24,'楊張彗':27}.items():
    for col in range(6, 34):
        excel_shifts[(name, col-5)] = normalize(ws2.cell(row=row, column=col).value)

for col in range(6, 34):
    day = col - 5
    v21 = normalize(ws2.cell(row=21, column=col).value)
    v22 = normalize(ws2.cell(row=22, column=col).value)
    excel_shifts[('蔡詠安', day)] = v21 if v21 != 'OFF' else v22

for day in range(1, 29):
    excel_shifts[('陳倩儀', day)] = normalize(ws2.cell(row=30, column=day+5).value)
    excel_shifts[('陳濬宏', day)] = 'D' if day == 28 else 'OFF'

with app.test_client() as c:
    sys_data = c.get('/api/schedule?year=2026&month=2').get_json()['data']

sys_shifts = {}
for row in sys_data:
    name = row['employee']['name']
    for d_str, shift in row['shifts'].items():
        sys_shifts[(name, int(d_str))] = shift

all_names = list(set(k[0] for k in excel_shifts.keys()))
sched_ok = 0
sched_fail = 0
for name in sorted(all_names):
    mismatches = []
    for day in range(1, 29):
        ex = excel_shifts.get((name, day), 'OFF')
        sy = sys_shifts.get((name, day), 'OFF')
        if ex != sy:
            mismatches.append(f'd{day}')
    if mismatches:
        check(f'{name} 差異: {mismatches}', False)
        sched_fail += 1
    else:
        sched_ok += 1

print(f'  排班正確: {sched_ok}/{sched_ok + sched_fail}人')

print()
print('=' * 60)
print('STEP 4: 薪資金額比對')
print('=' * 60)

excel_pay = {
    '施秉綬': 63582, '陳惠美': 21736, '王明珠': 39884, '何金艷': 52126,
    '全蕙萍': 56796, '衡怡紅': 30280, '藍茂禎': 72534,
    '邱素蘭': 47520,
    '林淑真': 43470, '歐南蘭': 41850, '沈佳榕': 35670, '鍾宏枝': 10400,
    '張瀞文': 19470, '王柏欣': 49696, '楊張彗': 62419, '陳倩儀': 16800,
}

with app.test_client() as c:
    salary_data = c.get('/api/salary/calculate?year=2026&month=2').get_json()['data']

for s in salary_data:
    name = s['employee']['name']
    sys_pay = s['actual_pay']
    ex_pay = excel_pay.get(name)
    if ex_pay is not None:
        check(f'{name}: 系統={sys_pay:,} Excel={ex_pay:,}', sys_pay == ex_pay)
    else:
        print(f'  SKIP {name}: 系統={sys_pay:,} (Excel無完整資料)')

print()
print('=' * 60)
print('STEP 5: 單人 Excel 匯出')
print('=' * 60)

with app.test_client() as c:
    r = c.get('/api/salary/export/1?year=2026&month=2')
    check(f'HTTP {r.status_code}', r.status_code == 200)

    import tempfile, os
    tmp = os.path.join(tempfile.gettempdir(), 'test_export.xlsx')
    with open(tmp, 'wb') as f:
        f.write(r.data)

    ewb = openpyxl.load_workbook(tmp)
    ews = ewb.active

    check('公司名稱', '愛立康' in str(ews['A1'].value or ''))
    check('員工姓名', '施秉綬' in str(ews['A4'].value or ''))

    # Find actual pay row
    actual_pay = None
    for row in range(1, ews.max_row + 1):
        val = ews.cell(row=row, column=1).value
        if val and '實發' in str(val):
            actual_pay = ews.cell(row=row, column=8).value
            break

    check(f'實發金額={actual_pay}', actual_pay == 63582)

    # Check earnings detail
    check(f'日班明細', '日班16天' in str(ews.cell(row=8, column=2).value or ''))
    check(f'夜班明細', '夜班5天' in str(ews.cell(row=9, column=2).value or ''))

    ewb.close()
    try:
        os.remove(tmp)
    except PermissionError:
        pass

print()
print('=' * 60)
print('STEP 6: 全體 Excel 匯出')
print('=' * 60)

with app.test_client() as c:
    r = c.get('/api/salary/export-all?year=2026&month=2')
    check(f'HTTP {r.status_code}', r.status_code == 200)

    if r.status_code == 200:
        tmp = os.path.join(tempfile.gettempdir(), 'test_all.xlsx')
        with open(tmp, 'wb') as f:
            f.write(r.data)

        ewb = openpyxl.load_workbook(tmp)
        check(f'Sheet數={len(ewb.sheetnames)}', len(ewb.sheetnames) == 19)

        # Spot check a few sheets
        for sheet_name in ['施秉綬', '何金艷', '王柏欣', '楊張彗']:
            if sheet_name in ewb.sheetnames:
                ws = ewb[sheet_name]
                pay = None
                for row in range(1, ws.max_row + 1):
                    val = ws.cell(row=row, column=1).value
                    if val and '實發' in str(val):
                        pay = ws.cell(row=row, column=8).value
                        break
                expected = excel_pay.get(sheet_name, 0)
                check(f'{sheet_name} 實發={pay} (預期{expected})', pay == expected)

        ewb.close()
        try:
            os.remove(tmp)
        except PermissionError:
            pass

print()
print('=' * 60)
if errors == 0:
    print('ALL PASSED - 全部驗證通過!')
else:
    print(f'FAILED - {errors} 項驗證失敗')
print('=' * 60)
